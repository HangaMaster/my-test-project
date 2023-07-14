from openpyxl.reader.excel import load_workbook
from ursina import *

# app = Ursina(borderless=False)

wb = load_workbook("map.xlsx", data_only=True)
ws = wb.worksheets[0]
cell = ws.cell(1, 1)

for row in range(10):
    for col in range(10):
        color_code = ws.cell(row + 1, col + 1).fill.start_color.index
        try:
            color = int(color_code[2:4], 16), int(color_code[4:6], 16), int(color_code[6:8], 16)
            print(color)
        except:
            print("Sin color")
app.run()
